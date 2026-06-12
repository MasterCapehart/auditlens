"""
AuditLens Excel Exporter — generates .xlsx with multiple sheets:
  - Summary     : severity counts + compliance pivot
  - Findings    : full findings table
  - By File     : risk score per file

Usage:
    auditlens scan ./project --format xlsx -o report.xlsx
"""

from __future__ import annotations

import os
from datetime import datetime
from typing import List

_SEVERITY_ORDER = {'CRITICAL': 0, 'HIGH': 1, 'MEDIUM': 2, 'LOW': 3}
_SEVERITY_FILL = {
    'CRITICAL': 'FFEF4444',
    'HIGH':     'FFF97316',
    'MEDIUM':   'FFEAB308',
    'LOW':      'FF3B82F6',
}


def generate_xlsx_report(
    findings: List[dict],
    scan_path: str,
    output_path: str = 'audit_report.xlsx',
) -> None:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import (
            Font, PatternFill, Alignment, Border, Side, GradientFill
        )
        from openpyxl.chart import BarChart, Reference
        from openpyxl.chart.series import DataPoint
        from openpyxl.utils import get_column_letter
    except ImportError:
        print(
            '\033[91m[AuditLens]\033[0m openpyxl no instalado.\n'
            'Instala con: pip install openpyxl --break-system-packages'
        )
        return

    wb = Workbook()
    now = datetime.now().strftime('%Y-%m-%d %H:%M')

    # ── helpers ───────────────────────────────────────────────────────────────
    def _header_font():
        return Font(bold=True, color='FFFFFFFF', size=11)

    def _header_fill():
        return PatternFill('solid', fgColor='FF1E3A5F')

    def _thin_border():
        s = Side(style='thin', color='FFE2E8F0')
        return Border(left=s, right=s, top=s, bottom=s)

    def _auto_width(ws):
        for col in ws.columns:
            max_len = max((len(str(c.value or '')) for c in col), default=0)
            ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 60)

    # ── Sheet 1: Summary ──────────────────────────────────────────────────────
    ws_sum = wb.active
    ws_sum.title = 'Summary'

    ws_sum['A1'] = '🛡️ AuditLens Security Report'
    ws_sum['A1'].font = Font(bold=True, size=16, color='FF1E3A5F')
    ws_sum['A2'] = f'Scan path: {scan_path}'
    ws_sum['A2'].font = Font(italic=True, color='FF64748B')
    ws_sum['A3'] = f'Generated: {now}'
    ws_sum['A3'].font = Font(italic=True, color='FF64748B')
    ws_sum.append([])

    counts = {'CRITICAL': 0, 'HIGH': 0, 'MEDIUM': 0, 'LOW': 0}
    for f in findings:
        sev = f.get('severity', 'LOW').upper()
        if sev in counts:
            counts[sev] += 1

    ws_sum.append(['Severity', 'Count'])
    for cell in ws_sum[ws_sum.max_row]:
        cell.font = _header_font()
        cell.fill = _header_fill()

    sev_start_row = ws_sum.max_row + 1
    for sev, cnt in counts.items():
        row = ws_sum.max_row + 1
        ws_sum.append([sev, cnt])
        fill_color = _SEVERITY_FILL.get(sev, 'FFCCCCCC')
        ws_sum.cell(row, 1).fill = PatternFill('solid', fgColor=fill_color)
        ws_sum.cell(row, 1).font = Font(bold=True, color='FFFFFFFF')
    sev_end_row = ws_sum.max_row

    ws_sum.append([])
    ws_sum.append(['Total Findings', len(findings)])
    ws_sum.cell(ws_sum.max_row, 1).font = Font(bold=True)

    # Bar chart for severity
    chart = BarChart()
    chart.type = 'col'
    chart.title = 'Findings by Severity'
    chart.y_axis.title = 'Count'
    chart.x_axis.title = 'Severity'
    chart.style = 10
    chart.width = 18
    chart.height = 12

    data_ref = Reference(ws_sum, min_col=2, min_row=sev_start_row - 1, max_row=sev_end_row)
    cats_ref = Reference(ws_sum, min_col=1, min_row=sev_start_row, max_row=sev_end_row)
    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(cats_ref)
    ws_sum.add_chart(chart, 'D5')

    _auto_width(ws_sum)

    # ── Sheet 2: Findings ─────────────────────────────────────────────────────
    ws_find = wb.create_sheet('Findings')
    headers = ['Severity', 'Rule ID', 'Name', 'File', 'Line', 'Description', 'Compliance']
    ws_find.append(headers)
    for cell in ws_find[1]:
        cell.font = _header_font()
        cell.fill = _header_fill()
        cell.alignment = Alignment(wrap_text=True)

    sorted_findings = sorted(findings, key=lambda x: _SEVERITY_ORDER.get(x.get('severity', 'LOW').upper(), 3))
    for f in sorted_findings:
        sev = f.get('severity', 'LOW').upper()
        file_short = '/'.join(f.get('file', '').split('/')[-3:])
        ws_find.append([
            sev,
            f.get('rule_id', ''),
            f.get('name', ''),
            file_short,
            f.get('line', ''),
            f.get('description', '')[:200],
            ', '.join(f.get('compliance', [])),
        ])
        row = ws_find.max_row
        fill_color = _SEVERITY_FILL.get(sev, 'FFCCCCCC')
        ws_find.cell(row, 1).fill = PatternFill('solid', fgColor=fill_color)
        ws_find.cell(row, 1).font = Font(bold=True, color='FFFFFFFF')
        for col in range(1, 8):
            ws_find.cell(row, col).border = _thin_border()

    ws_find.row_dimensions[1].height = 22
    _auto_width(ws_find)

    # ── Sheet 3: By File ──────────────────────────────────────────────────────
    ws_file = wb.create_sheet('By File')
    ws_file.append(['File', 'Total', 'Critical', 'High', 'Medium', 'Low', 'Risk Score'])
    for cell in ws_file[1]:
        cell.font = _header_font()
        cell.fill = _header_fill()

    file_stats: dict = {}
    for f in findings:
        fp = f.get('file', 'unknown')
        sev = f.get('severity', 'LOW').upper()
        if fp not in file_stats:
            file_stats[fp] = {'CRITICAL': 0, 'HIGH': 0, 'MEDIUM': 0, 'LOW': 0}
        if sev in file_stats[fp]:
            file_stats[fp][sev] += 1

    def _risk(stats):
        return stats['CRITICAL'] * 8 + stats['HIGH'] * 4 + stats['MEDIUM'] * 2 + stats['LOW']

    for fp, stats in sorted(file_stats.items(), key=lambda x: -_risk(x[1])):
        total_file = sum(stats.values())
        ws_file.append([
            fp,
            total_file,
            stats['CRITICAL'],
            stats['HIGH'],
            stats['MEDIUM'],
            stats['LOW'],
            _risk(stats),
        ])

    _auto_width(ws_file)

    wb.save(output_path)
    abs_path = os.path.abspath(output_path)
    print(f'\033[92m[AuditLens]\033[0m Reporte Excel guardado: \033[1m{abs_path}\033[0m')
